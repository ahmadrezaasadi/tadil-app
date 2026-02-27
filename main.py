# ====================کتابخانه های مورد نیاز===========
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.spinner import Spinner, SpinnerOption
from kivy.uix.togglebutton import ToggleButton
from kivy.core.window import Window
from kivy.properties import NumericProperty
from kivy.metrics import dp
from kivy.clock import Clock

# ================== کتابخانه‌های اکسل ==================
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("برای خروجی اکسل، کتابخانه openpyxl را نصب کنید: pip install openpyxl")

# ================== کتابخانه ذخیره فایل ==================
try:
    from plyer import filechooser
    PLYER_AVAILABLE = True
except ImportError:
    PLYER_AVAILABLE = False
    print("برای ذخیره فایل، کتابخانه plyer را نصب کنید: pip install plyer")

import os

# ================== تنظیم فونت فارسی ==================
from kivy.core.text import LabelBase

font_path = os.path.join(os.path.dirname(__file__), 'Vazir.ttf')
if os.path.exists(font_path):
    LabelBase.register(name='Vazir', fn_regular=font_path)
    print("فونت Vazir با موفقیت بارگذاری شد.")
else:
    print("خطا: فایل Vazir.ttf در کنار برنامه یافت نشد. از فونت پیش‌فرض استفاده می‌شود.")

# ================== کلاس سفارشی برای آیتم‌های Spinner ==================
class VazirSpinnerOption(SpinnerOption):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.font_name = 'Vazir'

# ================== توابع کمکی برای تبدیل اعداد فارسی به انگلیسی ==================
persian_digits = '۰۱۲۳۴۵۶۷۸۹'
english_digits = '0123456789'

def to_english_numbers(text):
    """تبدیل اعداد فارسی به انگلیسی در یک رشته"""
    if not text:
        return text
    trans_table = str.maketrans(persian_digits, english_digits)
    return text.translate(trans_table)

def to_persian_numbers(text):
    """تبدیل اعداد انگلیسی به فارسی (برای نمایش)"""
    if not text:
        return text
    trans_table = str.maketrans(english_digits, persian_digits)
    return text.translate(trans_table)

# ================== تابع اصلاح متن فارسی (با popup در صورت نداشتن کتابخانه) ==================
_fa_warning_shown = False  # فلگ برای نمایش یک بار پیام

def fa(text):
    """تبدیل متن فارسی به شکل درست برای نمایش (چسباندن حروف و راست‌به‌چپ کردن)"""
    if not text:
        return text
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)
    except ImportError:
        global _fa_warning_shown
        if not _fa_warning_shown:
            try:
                app = App.get_running_app()
                if app and app.root:
                    popup = Popup(
                        title='اخطار',
                        title_font='Vazir',
                        content=Label(
                            text='برای نمایش بهتر فارسی، کتابخانه‌های arabic_reshaper و python-bidi باید نصب شوند.\nممکن است متون به درستی نمایش داده نشوند.',
                            font_name='Vazir'
                        ),
                        size_hint=(0.8, 0.4)
                    )
                    popup.open()
            except:
                pass
            print("برای نمایش بهتر فارسی، کتابخانه‌های 'arabic_reshaper' و 'python-bidi==0.4.2' را نصب کنید.")
            _fa_warning_shown = True
        return text

# ================== تابع فرمت عدد فارسی ==================
def format_persian_number(num):
    try:
        s = f"{float(num):,.0f}"
        return to_persian_numbers(s)
    except:
        return "۰"

# ================== داده‌های ثابت ==================
Q_DATA = {
    "سازه‌وابنیه": {1: 0.74, 2: 0.74},
    "پایپینگ یا لوله‌کشی ": {
        3: 1, 4: 1, 5: 0.85, 6: 0.85, 7: 1,
        8: 0.85, 9: 0.90, 10: 0.85, 11: 0.85,
        12: 0.60, 13: 0.90
    },
    "تجهیزات مکانیک ثابت ": {
        14: 0.9, 15: 0.8, 16: 0.8, 17: 0.85, 18: 0.85
    },
    "تجهیزات مکانیکی دوّار ": {
        19: 0.85, 20: 0.85, 21: 0.90, 22: 0.85
    },
    "برق، ابزار دقیق و مخابرات": {
        23: 1, 24: 0.92, 25: 0.85, 26: 0.95, 27: 0.65,
        28: 1, 29: 1, 30: 0.95, 31: 0.90, 32: 0.85,
        33: 0.90, 34: 0.85, 35: 0.85, 36: 1, 37: 0.65,
        38: 0.95, 39: 0.90, 40: 0.90, 41: 0.84, 42: 0.92,
        43: 0.78, 44: 1, 45: 1, 46: 0.81, 47: 1,
        48: 0.85, 49: 1, 50: 0.90, 51: 0.85, 52: 0.90,
        53: 0.90, 54: 1, 55: 1, 56: 0.85, 57: 1,
        58: 0.95, 59: 0.78, 60: 0.85, 61: 1, 62: 1,
        63: 1, 64: 1, 65: 1, 66: 1, 67: 0.85, 68: 0.85,
        69: 0.90, 70: 0.74, 71: 0.70, 72: 0.85, 73: 0.80
    },
    "رنگ، عایق و مواد شیمیایی": {
        74: 0.90, 75: 0.50, 76: 0.65, 77: 0.50, 78: 0.85
    },
    "پکیج": {
        79: 0.74, 80: 0.90, 81: 0.90, 82: 0.90,
        83: 0.85, 84: 0.85, 85: 0.84, 86: 0.85
    }
}

Q_DESC = {
    1: "سازه‌ها، تکیه‌گاه‌ها، نرده‌ها، رک‌های لوله، سازه‌های فلزی و گریتینگ‌ها",
    2: "تمام انواع الکترودهای جوشکاری",
    3: "تمام انواع لوله‌های فولادی",
    4: "تمام انواع لوله‌های مسی",
    5: "اتصالات و فلنج‌ها، اسپیسرها و اسپک‌بلایندها",
    6: "تمام انواع شیرآلات پنوماتیک، برقی و کاهنده فشار",
    7: "لوله‌ها، اتصالات و شیرآلات پلی‌اتیلن",
    8: "صافی‌ها و فیلترها",
    9: "تمام انواع پیگ‌ها",
    10: "صداگیرها",
    11: "تله‌های بخار",
    12: "تمام انواع گسکت‌ها و اورینگ‌ها",
    13: "شعله‌گیرها",
    14: "تمام انواع ورق‌های فولادی برای ساخت مخازن",
    15: "درام‌ها، راکتورها، مخازن، خشک‌کن‌ها، برج‌ها و دِاراتورها",
    16: "برج‌ها و جداکننده‌های نفت و گاز",
    17: "تمام انواع مبدل‌های حرارتی (کندانسورها و مبدل‌ها)",
    18: "تمام انواع جرثقیل‌ها شامل سقفی، دروازه‌ای و غیره",
    19: "تمام انواع پمپ‌ها و سیل‌های مکانیکی",
    20: "تمام انواع کمپرسورها، توربوکمپرسورها و توربواکسپندرها",
    21: "کمپرسورهای تبرید",
    22: "میکسرها و اجکتورها",
    23: "ترانسفورماتور",
    24: "توربین",
    25: "تمام انواع موتورهای الکتریکی",
    26: "تابلوهای برق، تابلوهای کنترل و پنل‌ها",
    27: "جعبه‌های تقسیم، ترمینال‌ها و جعبه‌های توزیع صنعتی",
    28: "PT و CT",
    29: "تمام انواع کلیدهای قدرت",
    30: "تمام انواع رله‌ها، کنتاکتورها و فیوزها",
    31: "UPS و شارژرهای صنعتی",
    32: "تمام انواع باتری‌های صنعتی",
    33: "کابل‌های قدرت فشار متوسط و قوی",
    34: "کابل‌های ردیابی حرارتی برقی",
    35: "کابل‌های فشار ضعیف، ابزار دقیق، کنترل، F&G، مخابرات و شبکه",
    36: "کابل‌های فیبر نوری",
    37: "سینی‌های کابل، نردبان کابل و لوله‌های برق",
    38: "ابزار اندازه‌گیری جریان و ولتاژ",
    39: "تجهیزات روشنایی صنعتی",
    40: "سیستم‌های اعلام حریق",
    41: "تجهیزات آتش‌نشانی",
    42: "تمام انواع ژنراتورها",
    43: "بخاری‌های برقی",
    44: "سیستم‌های کنترل و ایمنی (FGS, ESD, DCS, PLC, PCS)",
    45: "سیستم‌های SCADA",
    46: "تجهیزات هیدرولیک و HPU",
    47: "پنل‌های خورشیدی",
    48: "تمام انواع آنالایزرها",
    49: "کامپیوترهای صنعتی و موارد مرتبط",
    50: "تمام انواع فشارسنج‌ها، دما سنج‌ها، سطح سنج‌ها و فلومترها",
    51: "تمام انواع ترانسمیترهای فشار، دما، سطح و جریان",
    52: "اوریفیس",
    53: "کنتورها و رگولاتورها",
    54: "تمام انواع خازن‌های صنعتی و بانک‌های خازنی",
    55: "راکتورهای قدرت",
    56: "دمنده‌ها، فن‌های صنعتی و کولرهای هوایی",
    57: "تمام انواع سیستم‌های ارتباطی، رادیویی و تلفنی",
    58: "تمام انواع سیستم‌های پیجینگ، آژیرها، زنگ‌ها و بوق‌ها",
    59: "تمام انواع دوربین‌ها و سیستم‌های حفاظت الکترونیکی و پیرامونی",
    60: "بالابرها و نوار نقاله‌ها",
    61: "داکت باس",
    62: "باسبار",
    63: "SPD (صاعقه‌گیر / برقگیر)",
    64: "تجهیزات سیستم ارت (میله‌ها، صفحات مسی، نوارها و غیره)",
    65: "تجهیزات حفاظت کاتدی",
    66: "تجهیزات پایش خوردگی (کوپن‌ها و پروب‌ها)",
    67: "دسوپرهیتر",
    68: "تجهیزات سرچاهی",
    69: "لوله‌های حفاری",
    70: "مته‌های حفاری",
    71: "گل حفاری",
    72: "سیمان حفاری",
    73: "بویه شناور SPM/SBM",
    74: "تمام انواع رنگ‌ها، پوشش‌ها، ماستیک‌ها، آسترها و سندبلاست",
    75: "عایق حرارتی و پوشش‌های ضد حریق (پلی‌اورتان)",
    76: "عایق رطوبتی (بر پایه روغن، قطران، بیتوسیل)",
    77: "عایق پلیمری",
    78: "تمام انواع مواد شیمیایی مورد استفاده در صنایع نفت، گاز و پتروشیمی",
    79: "فلر",
    80: "بویلرها و مولدهای بخار",
    81: "مشعل‌های بویلر",
    82: "سیستم‌ها و اجزای HVAC",
    83: "واحدهای شیرین‌سازی آب",
    84: "اندازه‌گیری و پروینگ",
    85: "پکیج تولید نیتروژن",
    86: "پکیج تولید هوا"
}

Q_ROW_INFO = {
    1: "فصل ۹ - رشته ابنیه",
    2: "فصل ۹ - رشته ابنیه",
    3: "فصل ۱۶ - رشته خطوط انتقال آب",
    4: "فصل ۶ - رشته تأسیسات مکانیکی",
    5: "فصل ۱۶ - رشته تجهیزات آب و فاضلاب",
    6: "فصل ۷ - رشته تأسیسات مکانیکی",
    7: "فصل ۱۴ - رشته شبکه توزیع آب",
    8: "فصل ۱۱ - رشته تأسیسات مکانیکی",
    9: "فصل ۸ - رشته تأسیسات مکانیکی",
    10: "فصل ۹ - رشته تأسیسات مکانیکی",
    11: "فصل ۱۱ - رشته تأسیسات مکانیکی",
    12: "فصل ۲۵ - رشته تأسیسات مکانیکی",
    13: "فصل ۸ - رشته تأسیسات مکانیکی",
    14: "فصل ۱۰ - رشته راه، راه‌آهن و باند فرودگاه",
    15: "فصل ۳۳ - رشته تأسیسات مکانیکی",
    16: "فصل ۳۳ - رشته تأسیسات مکانیکی",
    17: "فصل ۹ - رشته تجهیزات آب و فاضلاب",
    18: "فصل ۱۲ - رشته تجهیزات آب و فاضلاب",
    19: "فصل ۱ - رشته تجهیزات آب و فاضلاب",
    20: "فصل ۴ - رشته تجهیزات آب و فاضلاب",
    21: "فصل ۲۷ - رشته تأسیسات مکانیکی",
    22: "فصل ۲ - رشته تجهیزات آب و فاضلاب",
    23: "فصل ۲ - رشته انتقال و توزیع نیرو",
    24: "فصل ۱۷ - رشته تأسیسات برقی",
    25: "فصل ۱۳ - رشته تجهیزات آب و فاضلاب",
    26: "فصل ۱۴ - رشته تأسیسات برقی",
    27: "فصل ۲۸ - رشته تأسیسات برقی",
    28: "فصل ۱۳ - رشته توزیع نیرو",
    29: "فصل ۵ - رشته انتقال و توزیع نیرو",
    30: "فصل ۱۴ - رشته تأسیسات برقی",
    31: "فصل ۲۸ - رشته انتقال و توزیع نیرو",
    32: "فصل ۳۰ - رشته انتقال و توزیع نیرو",
    33: "فصل ۲ - رشته انتقال و توزیع نیرو زیرزمینی",
    34: "فصل ۷ - رشته تأسیسات برقی",
    35: "فصل ۷ - رشته تأسیسات برقی",
    36: "فصل ۷ - رشته انتقال و توزیع نیرو زیرزمینی",
    37: "فصل ۲۸ - رشته تأسیسات برقی",
    38: "فصل ۱۵ - رشته تأسیسات برقی",
    39: "فصل ۵ - رشته تأسیسات برقی",
    40: "فصل ۲۶ - رشته تأسیسات برقی",
    41: "رشته تأسیسات مکانیکی",
    42: "فصل ۱۷ - رشته تأسیسات برقی",
    43: "رشته تأسیسات برقی",
    44: "فصل ۱۷ - رشته توزیع نیرو",
    45: "فصل ۱۷ - رشته توزیع نیرو",
    46: "میانگین شاخص رشته‌های تأسیسات مکانیکی و برقی",
    47: "فصل ۳۵ - رشته تأسیسات برقی",
    48: "فصل ۳۳ - رشته تجهیزات آب و فاضلاب",
    49: "فصل ۲۴ - رشته انتقال و توزیع نیرو",
    50: "فصل ۱۵ - رشته تأسیسات مکانیکی",
    51: "فصل ۳۱ - رشته تجهیزات آب و فاضلاب",
    52: "فصل ۱۵ - رشته تأسیسات مکانیکی",
    53: "فصل ۱۵ - رشته تأسیسات مکانیکی",
    54: "فصل ۱۶ - رشته انتقال و توزیع نیرو",
    55: "فصل ۳ - رشته انتقال و توزیع نیرو",
    56: "فصل ۴ - رشته تجهیزات آب و فاضلاب",
    57: "فصل ۲۶ - رشته انتقال و توزیع نیرو",
    58: "فصل ۲۷ - رشته تأسیسات برقی",
    59: "رشته تأسیسات برقی",
    60: "فصل ۱۳ - رشته تجهیزات آب و فاضلاب",
    61: "فصل ۱۷ - رشته انتقال و توزیع نیرو",
    62: "فصل ۲۵ - رشته توزیع نیرو",
    63: "فصل ۲۱ - رشته توزیع نیرو",
    64: "فصل ۱۸ - رشته انتقال و توزیع نیرو",
    65: "فصل ۵ - رشته بهره‌برداری و نگهداری تأسیسات آب آشامیدنی",
    66: "فصل ۵ - رشته بهره‌برداری و نگهداری تأسیسات آب آشامیدنی",
    67: "فصل ۷ - رشته تأسیسات مکانیکی",
    68: "فصل ۷ - رشته تأسیسات مکانیکی",
    69: "فصل ۵ - رشته چاه‌ها",
    70: "فصل ۹ - رشته ابنیه",
    71: "فصل ۵ - رشته راه، راه‌آهن و باند فرودگاه",
    72: "فصل ۸ - رشته ابنیه",
    73: "فصل ۱۲ - رشته کارهای دریایی و ساحلی",
    74: "فصل ۱۶ - رشته راهداری",
    75: "فصل ۱۴ - رشته ابنیه",
    76: "فصل ۱۳ - رشته ابنیه",
    77: "فصل ۱۴ - رشته ابنیه",
    78: "فصل ۲۴ - رشته تجهیزات آب و فاضلاب",
    79: "فصل ۹ - رشته ابنیه",
    80: "فصل ۱۳ - رشته تأسیسات مکانیکی",
    81: "فصل ۱۴ - رشته تأسیسات مکانیکی",
    82: "فصل ۲۷ - رشته تأسیسات مکانیکی",
    83: "رشته تجهیزات آب و فاضلاب",
    84: "فصل ۳۱ - رشته تجهیزات آب و فاضلاب",
    85: "رشته تأسیسات مکانیکی",
    86: "فصل ۴ - رشته تجهیزات آب و فاضلاب"
}

BILL_OF_QUANTITIES_LIB = {
    "گروه ۱": {
        "Lists": [
            "۱- خطوط لوله انتقال نفت و گاز بین‌شهری",
            "۲- خطوط لوله گاز شهری (به استثنای بخش پلی‌اتیلن)",
            "۳- خطوط لوله کمربندی و خطوط تغذیه نفت و گاز",
            "۴- تعمیر خطوط لوله کمربندی، خطوط تغذیه و شبکه گاز",
            "۵- خطوط لوله جریانی رو زمینی نفت و گاز",
            "۶- گازرسانی به صنایع"
        ],
        "Corresponding Index": "شاخص فصل ۴، عملیات لوله گذاری با لوله های فولادی اتصال جوشی فهرست بهای پایه رشته خطوط انتقال آب"
    },
    "گروه ۲": {
        "Lists": [
            "فهرست بهای عملیات ساختمانی صنعتی نفت و گاز و پتروشیمی"
        ],
        "Corresponding Index": "شاخص رشته ای فهرست بهای پایه رشته ابنیه"
    },
    "گروه ۳": {
        "Lists": [
            "بخش لوله‌های پلی‌اتیلن در فهرست بهای رشته خطوط لوله گاز شهری"
        ],
        "Corresponding Index": "شاخص فصل ۴، عملیات لوله‌گذاری با لوله پلی‌اتیلن فهرست بهای پایه رشته شبکه توزیع آب"
    },
    "گروه ۴": {
        "Lists": [
            "۱- نصب پالایشگاه‌های نفت و گاز، واحدهای پتروشیمی و واحدهای تفکیک مایعات گازی NGL",
            "۲- نصب تلمبه خانه های نفت و انبارهای نفت منطقه‌ای",
            "۳- نصب واحدهای بهره برداری نفت و گاز و ایستگاه‌های تراکم گاز",
            "۴- نصب واحدهای سرچاهی نفت و گاز و چندراهه ها",
            "۵-  تعمیرات پالایشگاه‌",
            "۶- تعمیرات تاسیسات ساحلی و فراساحل جزایر در صنعت نفت "
        ],
        "Subgroups": {
            "لوله‌کشی": {
                "Corresponding Index": "شاخص فصل ۳۵، کارهای دستمزدی فهرست بهای پایه تأسیسات مکانیکی (۷۰٪) + شاخص فصل ۳، عملیات خاکی با ماشین فهرست بهای پایه ابنیه (۳۰٪)"
            },
            "نصب تجهیزات، اسکلت فلزی و رنگ‌": {
                "Corresponding Index": "شاخص فصل ۳۵، کارهای دستمزدی فهرست بهای پایه تأسیسات مکانیکی (۴۵٪) + شاخص فصل ۳، کارهای خاکی با ماشین فهرست بهای پایه ابنیه (۵۵٪)"
            },
            "مخازن ذخیره": {
                "Corresponding Index": "شاخص فصل ۳۵، کارهای دستمزدی فهرست بهای پایه تأسیسات مکانیکی (۶۰٪) + شاخص فصل ۳، کارهای خاکی با ماشین فهرست بهای پایه ابنیه (۴۰٪)"
            },
            "عایق‌، کارهای برقی و ابزار دقیق": {
                "Corresponding Index": "شاخص فصل ۳۵، کارهای دستمزدی فهرست بهای پایه تأسیسات مکانیکی (۹۰٪) + شاخص فصل ۳، کارهای خاکی با ماشین فهرست بهای پایه ابنیه (۱۰٪)"
            }
        }
    }
}

# ================== توابع محاسباتی ==================
def calc_alpha_buy(q, I_em, I_bm, I_ew, I_bw, delay_factor):
    if I_bm == 0 or I_bw == 0:
        raise ValueError(fa("شاخص مبنا نمی‌تواند صفر باشد"))
    factor = ((I_em / I_bm) - (1 - q) * (I_ew / I_bw)) / q
    return round(delay_factor * (factor - 1), 3)

def calc_alpha_simple(exec_index, base_index, delay_factor):
    if base_index == 0:
        raise ValueError(fa("شاخص مبنا نمی‌تواند صفر باشد"))
    return round(delay_factor * ((exec_index / base_index) - 1), 3)

# ================== صفحه شروع ==================
class StartScreen(BoxLayout):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.app = app
        self.orientation = 'vertical'
        self.padding = dp(20)
        self.spacing = dp(15)

        title = Label(
            text=fa("محاسبه تعدیل"),
            font_size=dp(20),
            bold=True,
            color=(0.2, 0.4, 0.8, 1),
            size_hint=(1, 0.3),
            font_name='Vazir'
        )
        self.add_widget(title)

        input_layout = BoxLayout(orientation='vertical', spacing=dp(8), size_hint=(1, 0.4))
        input_label = Label(
            text=fa("تعداد صورت‌وضعیت‌ها را وارد کنید:"),
            font_size=dp(14),
            font_name='Vazir'
        )
        input_layout.add_widget(input_label)

        self.statement_input = TextInput(
            multiline=False,
            input_filter='int',
            font_size=dp(16),
            halign='center',
            size_hint=(1, None),
            height=dp(60),
            padding=(dp(15), dp(15)),
            font_name='Vazir',
            hint_text=fa("مثال ۲")
        )
        input_layout.add_widget(self.statement_input)
        self.add_widget(input_layout)

        start_btn = Button(
            text=fa("شروع"),
            font_size=dp(16),
            background_color=(0.2, 0.6, 0.2, 1),
            size_hint=(1, None),
            height=dp(60),
            font_name='Vazir'
        )
        start_btn.bind(on_press=self.start_calculation)
        self.add_widget(start_btn)

    def start_calculation(self, instance):
        try:
            text = to_english_numbers(self.statement_input.text)
            count = int(text)
            if count > 0:
                self.app.statement_count = count
                self.app.statement_results = []
                self.app.statement_details = []
                self.app.base_indices = {}
                self.app.show_boq_selection()
            else:
                self.show_error(fa("عدد باید بزرگتر از صفر باشد"))
        except:
            self.show_error(fa("لطفاً یک عدد معتبر وارد کنید"))

    def show_error(self, message):
        popup = Popup(
            title=fa('خطا'),
            title_font='Vazir',
            content=Label(text=message, font_name='Vazir'),
            size_hint=(0.8, 0.4)
        )
        popup.open()

# ================== صفحه انتخاب BOQ ==================
class BOQSelectionScreen(BoxLayout):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.app = app
        self.orientation = 'vertical'
        self.padding = dp(15)
        self.spacing = dp(12)

        title = Label(
            text=fa("انتخاب صورت وضعیت"),
            font_size=dp(18),
            bold=True,
            size_hint=(1, 0.1),
            font_name='Vazir'
        )
        self.add_widget(title)

        group_layout = BoxLayout(orientation='vertical', spacing=dp(4), size_hint=(1, 0.15))
        group_label = Label(
            text=fa("گروه اصلی:"),
            font_size=dp(14),
            size_hint=(1, None),
            height=dp(30),
            font_name='Vazir'
        )
        group_layout.add_widget(group_label)

        self.boq_group_map = {fa(key): key for key in BILL_OF_QUANTITIES_LIB.keys()}
        self.group_spinner = Spinner(
            text=fa('انتخاب گروه'),
            values=list(self.boq_group_map.keys()),
            font_size=dp(14),
            size_hint=(1, None),
            height=dp(60),
            background_color=(0.9, 0.9, 0.9, 1),
            font_name='Vazir',
            option_cls=VazirSpinnerOption
        )
        self.group_spinner.bind(text=self.on_group_select)
        group_layout.add_widget(self.group_spinner)
        self.add_widget(group_layout)

        self.lists_label = Label(
            text="",
            font_size=dp(11),
            color=(0, 0.6, 0, 1),
            size_hint=(1, 0.35),
            text_size=(Window.width - dp(30), None),
            halign='left',
            valign='top',
            font_name='Vazir'
        )
        lists_scroll = ScrollView(size_hint=(1, 0.35))
        lists_scroll.add_widget(self.lists_label)
        self.add_widget(lists_scroll)

        self.subgroup_layout = BoxLayout(orientation='vertical', spacing=dp(4), size_hint=(1, 0.1))
        self.subgroup_layout.opacity = 0

        subgroup_label = Label(
            text=fa("زیرگروه:"),
            font_size=dp(14),
            size_hint=(1, None),
            height=dp(30),
            font_name='Vazir'
        )
        self.subgroup_layout.add_widget(subgroup_label)

        self.subgroup_spinner = Spinner(
            text=fa('انتخاب زیرگروه'),
            font_size=dp(14),
            size_hint=(1, None),
            height=dp(60),
            background_color=(0.9, 0.9, 0.9, 1),
            font_name='Vazir',
            option_cls=VazirSpinnerOption
        )
        self.subgroup_spinner.bind(text=self.on_subgroup_select)
        self.subgroup_layout.add_widget(self.subgroup_spinner)
        self.add_widget(self.subgroup_layout)

        self.index_label = Label(
            text="",
            font_size=dp(12),
            color=(0.1, 0.5, 0.1, 1),
            size_hint=(1, 0.1),
            text_size=(Window.width - dp(30), None),
            halign='left',
            valign='top',
            font_name='Vazir'
        )
        self.add_widget(self.index_label)

        btn_layout = BoxLayout(spacing=dp(15), size_hint=(1, 0.15))

        back_btn = Button(
            text=fa("بازگشت"),
            font_size=dp(14),
            background_color=(0.8, 0.3, 0.3, 1),
            size_hint=(0.5, None),
            height=dp(60),
            font_name='Vazir'
        )
        back_btn.bind(on_press=lambda x: self.app.show_start_screen())
        btn_layout.add_widget(back_btn)

        self.next_btn = Button(
            text=fa("بعدی"),
            font_size=dp(14),
            background_color=(0.2, 0.6, 0.2, 1),
            disabled=True,
            size_hint=(0.5, None),
            height=dp(60),
            font_name='Vazir'
        )
        self.next_btn.bind(on_press=self.proceed)
        btn_layout.add_widget(self.next_btn)

        self.add_widget(btn_layout)

        self.selected_info = {}

    def on_group_select(self, spinner, text):
        if text == fa('انتخاب گروه'):
            return
        original_group = self.boq_group_map.get(text)
        if not original_group:
            return
        group_name = original_group
        self.selected_info['group'] = group_name

        lists = BILL_OF_QUANTITIES_LIB[group_name]["Lists"]
        self.lists_label.text = "\n".join([fa(f"• {item}") for item in lists])

        if group_name == "گروه ۴":
            self.subgroup_layout.opacity = 1
            subgroups = list(BILL_OF_QUANTITIES_LIB[group_name]["Subgroups"].keys())
            self.subgroup_spinner.values = [fa(sg) for sg in subgroups]
            self.subgroup_spinner.text = fa('انتخاب زیرگروه')
            self.selected_info['subgroup'] = None
            self.index_label.text = ""
        else:
            self.subgroup_layout.opacity = 0
            self.selected_info['subgroup'] = None
            self.update_index()

    def on_subgroup_select(self, spinner, text):
        if text == fa('انتخاب زیرگروه'):
            return
        group_name = self.selected_info.get('group')
        if group_name and group_name == "گروه ۴":
            subgroups_orig = list(BILL_OF_QUANTITIES_LIB[group_name]["Subgroups"].keys())
            subgroups_display = [fa(sg) for sg in subgroups_orig]
            try:
                idx = subgroups_display.index(text)
                self.selected_info['subgroup'] = subgroups_orig[idx]
            except ValueError:
                self.selected_info['subgroup'] = None
        self.update_index()

    def update_index(self):
        group = self.selected_info.get('group')
        subgroup = self.selected_info.get('subgroup')

        if group == "گروه ۴" and subgroup:
            index = BILL_OF_QUANTITIES_LIB[group]["Subgroups"][subgroup]["Corresponding Index"]
            self.selected_info['value'] = index
            self.index_label.text = fa(f"شاخص متناظر:\n{index}")
            self.next_btn.disabled = False
        elif group and group != "گروه ۴":
            index = BILL_OF_QUANTITIES_LIB[group]["Corresponding Index"]
            self.selected_info['value'] = index
            self.index_label.text = fa(f"شاخص متناظر:\n{index}")
            self.next_btn.disabled = False
        else:
            self.next_btn.disabled = True

    def proceed(self, instance):
        if not self.next_btn.disabled:
            self.app.boq_info = self.selected_info
            self.app.show_q_selection()

# ================== صفحه انتخاب Q ==================
class QSelectionScreen(BoxLayout):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.app = app
        self.orientation = 'vertical'
        self.padding = dp(15)
        self.spacing = dp(12)
        self.row_map = {}

        title = Label(
            text=fa("انتخاب کالا"),
            font_size=dp(18),
            bold=True,
            size_hint=(1, 0.1),
            font_name='Vazir'
        )
        self.add_widget(title)

        group_layout = BoxLayout(orientation='vertical', spacing=dp(4), size_hint=(1, 0.15))
        group_label = Label(
            text=fa("گروه کالایی:"),
            font_size=dp(14),
            size_hint=(1, None),
            height=dp(30),
            font_name='Vazir'
        )
        group_layout.add_widget(group_label)

        self.q_group_map = {fa(key): key for key in Q_DATA.keys()}
        self.group_spinner = Spinner(
            text=fa('انتخاب گروه'),
            values=list(self.q_group_map.keys()),
            font_size=dp(14),
            size_hint=(1, None),
            height=dp(60),
            background_color=(0.9, 0.9, 0.9, 1),
            font_name='Vazir',
            option_cls=VazirSpinnerOption
        )
        self.group_spinner.bind(text=self.on_group_select)
        group_layout.add_widget(self.group_spinner)
        self.add_widget(group_layout)

        row_layout = BoxLayout(orientation='vertical', spacing=dp(4), size_hint=(1, 0.2))
        row_label = Label(
            text=fa("ردیف کالایی:"),
            font_size=dp(14),
            size_hint=(1, None),
            height=dp(30),
            font_name='Vazir'
        )
        row_layout.add_widget(row_label)

        self.row_spinner = Spinner(
            text=fa('انتخاب ردیف'),
            font_size=dp(14),
            size_hint=(1, None),
            height=dp(60),
            background_color=(0.9, 0.9, 0.9, 1),
            font_name='Vazir',
            option_cls=VazirSpinnerOption
        )
        self.row_spinner.bind(text=self.on_row_select)
        row_layout.add_widget(self.row_spinner)
        self.add_widget(row_layout)

        self.q_label = Label(
            text=fa("ضریب q = ---"),
            font_size=dp(16),
            bold=True,
            color=(0.2, 0.2, 0.8, 1),
            size_hint=(1, None),
            height=dp(40),
            font_name='Vazir'
        )
        self.add_widget(self.q_label)

        self.desc_label = Label(
            text="",
            font_size=dp(11),
            color=(0, 0.6, 0, 1),
            size_hint=(1, 0.25),
            text_size=(Window.width - dp(30), None),
            halign='left',
            valign='top',
            font_name='Vazir'
        )
        self.add_widget(self.desc_label)

        btn_layout = BoxLayout(spacing=dp(15), size_hint=(1, 0.15))

        back_btn = Button(
            text=fa("بازگشت"),
            font_size=dp(14),
            background_color=(0.8, 0.3, 0.3, 1),
            size_hint=(0.5, None),
            height=dp(60),
            font_name='Vazir'
        )
        back_btn.bind(on_press=lambda x: self.app.show_boq_selection())
        btn_layout.add_widget(back_btn)

        self.next_btn = Button(
            text=fa("بعدی"),
            font_size=dp(14),
            background_color=(0.2, 0.6, 0.2, 1),
            disabled=True,
            size_hint=(0.5, None),
            height=dp(60),
            font_name='Vazir'
        )
        self.next_btn.bind(on_press=self.proceed)
        btn_layout.add_widget(self.next_btn)

        self.add_widget(btn_layout)

        self.selected_info = {}

    def on_group_select(self, spinner, text):
        if text == fa('انتخاب گروه'):
            return
        original_group = self.q_group_map.get(text)
        if not original_group:
            return
        group_name = original_group
        self.selected_info['group'] = group_name

        self.row_map = {}
        row_values = []
        for row_id in Q_DATA[group_name].keys():
            desc = Q_DESC.get(row_id, "")
            short_desc = (desc[:35] + "...") if len(desc) > 35 else desc
            display_text = fa(f"{to_persian_numbers(str(row_id))}|{short_desc}")
            self.row_map[display_text] = row_id
            row_values.append(display_text)
        self.row_spinner.values = row_values
        self.row_spinner.text = fa('انتخاب ردیف')
        self.q_label.text = fa("ضریب q = ---")
        self.desc_label.text = ""
        self.next_btn.disabled = True

    def on_row_select(self, spinner, text):
        if text == fa('انتخاب ردیف'):
            return
        row_id = self.row_map.get(text)
        if row_id is None:
            popup = Popup(
                title=fa('خطا'),
                title_font='Vazir',
                content=Label(text=fa("ردیف نامعتبر است"), font_name='Vazir'),
                size_hint=(0.8, 0.4)
            )
            popup.open()
            return

        group_name = self.selected_info.get('group')
        if not group_name:
            popup = Popup(
                title=fa('خطا'),
                title_font='Vazir',
                content=Label(text=fa("گروه کالایی انتخاب نشده است"), font_name='Vazir'),
                size_hint=(0.8, 0.4)
            )
            popup.open()
            return

        if group_name not in Q_DATA:
            popup = Popup(
                title=fa('خطا'),
                title_font='Vazir',
                content=Label(text=fa(f"گروه '{group_name}' در داده‌ها موجود نیست"), font_name='Vazir'),
                size_hint=(0.8, 0.4)
            )
            popup.open()
            return

        if row_id not in Q_DATA[group_name]:
            popup = Popup(
                title=fa('خطا'),
                title_font='Vazir',
                content=Label(text=fa(f"ردیف {row_id} در گروه '{group_name}' موجود نیست"), font_name='Vazir'),
                size_hint=(0.8, 0.4)
            )
            popup.open()
            return

        q_value = Q_DATA[group_name][row_id]
        desc = Q_DESC.get(row_id, "")
        row_info = Q_ROW_INFO.get(row_id, "")

        self.q_label.text = fa(f"ضریب q = {q_value}")
        self.desc_label.text = fa(f"توضیحات: {desc}")
        self.selected_info = {
            'value': q_value,
            'row': row_id,
            'group': group_name,
            'row_info': row_info
        }
        self.next_btn.disabled = False

    def proceed(self, instance):
        if not self.next_btn.disabled:
            self.app.q_info = self.selected_info
            self.app.show_delay_selection()

# ================== صفحه انتخاب تأخیر ==================
class DelaySelectionScreen(BoxLayout):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.app = app
        self.orientation = 'vertical'
        self.padding = dp(20)
        self.spacing = dp(15)

        title = Label(
            text=fa("وضعیت تأخیر قرارداد"),
            font_size=dp(18),
            bold=True,
            size_hint=(1, 0.15),
            font_name='Vazir'
        )
        self.add_widget(title)

        instruction = Label(
            text=fa("وضعیت تأخیر قرارداد را انتخاب کنید:"),
            font_size=dp(14),
            size_hint=(1, 0.1),
            font_name='Vazir'
        )
        self.add_widget(instruction)

        self.delay_var = 1

        options_layout = BoxLayout(orientation='vertical', spacing=dp(12), size_hint=(1, 0.5))

        self.no_delay_btn = ToggleButton(
            text=fa("قرارداد بدون تأخیر (۱.۰)"),
            group='delay',
            state='down',
            font_size=dp(14),
            size_hint=(1, None),
            height=dp(60),
            font_name='Vazir'
        )
        self.no_delay_btn.bind(on_press=lambda x: setattr(self, 'delay_var', 1))
        options_layout.add_widget(self.no_delay_btn)

        self.excusable_btn = ToggleButton(
            text=fa("قرارداد با تأخیر قابل قبول (۰.۹۷۵)"),
            group='delay',
            font_size=dp(14),
            size_hint=(1, None),
            height=dp(60),
            font_name='Vazir'
        )
        self.excusable_btn.bind(on_press=lambda x: setattr(self, 'delay_var', 2))
        options_layout.add_widget(self.excusable_btn)

        self.non_excusable_btn = ToggleButton(
            text=fa("قرارداد با تأخیر غیرقابل قبول (۰.۹۵)"),
            group='delay',
            font_size=dp(14),
            size_hint=(1, None),
            height=dp(60),
            font_name='Vazir'
        )
        self.non_excusable_btn.bind(on_press=lambda x: setattr(self, 'delay_var', 3))
        options_layout.add_widget(self.non_excusable_btn)

        self.add_widget(options_layout)

        btn_layout = BoxLayout(spacing=dp(15), size_hint=(1, 0.2))

        back_btn = Button(
            text=fa("بازگشت"),
            font_size=dp(14),
            background_color=(0.8, 0.3, 0.3, 1),
            size_hint=(0.5, None),
            height=dp(60),
            font_name='Vazir'
        )
        back_btn.bind(on_press=lambda x: self.app.show_q_selection())
        btn_layout.add_widget(back_btn)

        next_btn = Button(
            text=fa("شروع محاسبه"),
            font_size=dp(14),
            background_color=(0.2, 0.6, 0.2, 1),
            size_hint=(0.5, None),
            height=dp(60),
            font_name='Vazir'
        )
        next_btn.bind(on_press=self.proceed)
        btn_layout.add_widget(next_btn)

        self.add_widget(btn_layout)

    def proceed(self, instance):
        if self.delay_var == 1:
            delay_factor = 1.0
        elif self.delay_var == 2:
            delay_factor = 0.975
        else:
            delay_factor = 0.95
        self.app.delay_factor = delay_factor
        self.app.show_calculation_screen()

# ================== صفحه اصلی محاسبات (با امکان مبالغ صفر) ==================
class CalculationScreen(BoxLayout):
    current_statement = NumericProperty(1)
    total_adjustment = NumericProperty(0)

    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.app = app
        self.orientation = 'vertical'
        self.padding = dp(10)
        self.spacing = dp(8)

        self.total_days_so_far = 0
        self.prev_alpha_buy = 0
        self.prev_alpha_exec = 0
        self.prev_alpha_work = 0
        self.Ibm_base = None
        self.Ibw_base = None
        self.exec_base_value = None
        self.work_base_value = None

        self.iem_inputs = []
        self.iew_inputs = []
        self.exec_inputs = []
        self.work_inputs = []
        self.days_inputs = []

        self.header = Label(
            text=fa(f"صورت‌وضعیت {self.current_statement}"),
            font_size=dp(18),
            bold=True,
            size_hint=(1, None),
            height=dp(40),
            font_name='Vazir'
        )
        self.add_widget(self.header)

        period_layout = BoxLayout(orientation='horizontal', spacing=dp(10), size_hint=(1, None), height=dp(50))
        period_layout.add_widget(Label(
            text=fa("تعداد دوره‌های سه‌ماهه"),
            font_size=dp(14),
            size_hint=(0.5, 1),
            font_name='Vazir'
        ))
        self.period_spinner = Spinner(
            text='1',
            values=['1', '2', '3', '4'],
            font_size=dp(14),
            size_hint=(0.5, 1),
            font_name='Vazir',
            option_cls=VazirSpinnerOption
        )
        self.period_spinner.bind(text=self.on_period_change)
        period_layout.add_widget(self.period_spinner)
        self.add_widget(period_layout)

        self.q_row_info = self.app.q_info.get('row_info', '')
        self.boq_index_info = self.app.boq_info.get('value', '')

        self.scroll_content = BoxLayout(orientation='vertical', spacing=dp(8), size_hint_y=None)
        self.scroll_content.bind(minimum_height=self.scroll_content.setter('height'))

        # ========== خرید ==========
        purchase_title = Label(
            text=fa("اطلاعات خرید"),
            font_size=dp(15),
            bold=True,
            size_hint=(1, None),
            height=dp(40),
            color=(0.3, 0.5, 0.2, 1),
            font_name='Vazir'
        )
        self.scroll_content.add_widget(purchase_title)

        self.scroll_content.add_widget(Label(
            text=fa("مبلغ خرید:"),
            font_size=dp(13),
            size_hint=(1, None),
            height=dp(25),
            font_name='Vazir'
        ))
        self.buy_amt_input = TextInput(
            hint_text=fa("مبلغ را وارد کنید (ریال) - صفر مجاز است"),
            input_filter='float',
            size_hint=(1, None),
            height=dp(50),
            font_size=dp(14),
            multiline=False,
            padding=(dp(10), dp(12)),
            font_name='Vazir'
        )
        self.buy_amt_input.bind(text=lambda *x: self.update_fields_state())
        self.scroll_content.add_widget(self.buy_amt_input)

        self.scroll_content.add_widget(Label(
            text=fa("ضریب q:"),
            font_size=dp(13),
            size_hint=(1, None),
            height=dp(25),
            font_name='Vazir'
        ))
        self.q_input = TextInput(
            text=str(self.app.q_info['value']),
            readonly=True,
            size_hint=(1, None),
            height=dp(50),
            font_size=dp(14),
            multiline=False,
            padding=(dp(10), dp(12)),
            font_name='Vazir'
        )
        self.scroll_content.add_widget(self.q_input)

        iem_label_layout = BoxLayout(orientation='vertical', spacing=dp(2), size_hint=(1, None), height=dp(60))
        iem_label = Label(
            text=fa("شاخص دوره اجرا:"),
            font_size=dp(13),
            size_hint=(1, 0.4),
            font_name='Vazir'
        )
        iem_label_layout.add_widget(iem_label)
        iem_help_label = Label(
            text=fa(self.q_row_info),
            font_size=dp(9),
            color=(0, 0.6, 0, 1),
            size_hint=(1, 0.6),
            text_size=(Window.width - dp(40), None),
            halign='left',
            valign='top',
            font_name='Vazir'
        )
        iem_label_layout.add_widget(iem_help_label)
        self.scroll_content.add_widget(iem_label_layout)

        self.iem_container = BoxLayout(orientation='vertical', size_hint=(1, None), spacing=dp(5))
        self.scroll_content.add_widget(self.iem_container)

        iew_label = Label(
            text=fa("شاخص دستمزدی اجرا:"),
            font_size=dp(13),
            size_hint=(1, None),
            height=dp(25),
            font_name='Vazir'
        )
        self.scroll_content.add_widget(iew_label)
        self.iew_container = BoxLayout(orientation='vertical', size_hint=(1, None), spacing=dp(5))
        self.scroll_content.add_widget(self.iew_container)

        self.scroll_content.add_widget(Label(
            text=fa("شاخص مبنا:"),
            font_size=dp(13),
            size_hint=(1, None),
            height=dp(25),
            font_name='Vazir'
        ))
        self.Ibm_input = TextInput(
            hint_text=fa("شاخص مبنا را وارد کنید"),
            input_filter='float',
            size_hint=(1, None),
            height=dp(50),
            font_size=dp(14),
            multiline=False,
            padding=(dp(10), dp(12)),
            font_name='Vazir'
        )
        self.scroll_content.add_widget(self.Ibm_input)

        self.scroll_content.add_widget(Label(
            text=fa("شاخص مبنای دستمزدی:"),
            font_size=dp(13),
            size_hint=(1, None),
            height=dp(25),
            font_name='Vazir'
        ))
        self.Ibw_input = TextInput(
            hint_text=fa("شاخص مبنای دستمزدی را وارد کنید"),
            input_filter='float',
            size_hint=(1, None),
            height=dp(50),
            font_size=dp(14),
            multiline=False,
            padding=(dp(10), dp(12)),
            font_name='Vazir'
        )
        self.scroll_content.add_widget(self.Ibw_input)

        # ========== اجرا ==========
        exec_title = Label(
            text=fa("اطلاعات اجرا"),
            font_size=dp(15),
            bold=True,
            size_hint=(1, None),
            height=dp(40),
            color=(0.3, 0.5, 0.2, 1),
            font_name='Vazir'
        )
        self.scroll_content.add_widget(exec_title)

        self.scroll_content.add_widget(Label(
            text=fa("مبلغ اجرا:"),
            font_size=dp(13),
            size_hint=(1, None),
            height=dp(25),
            font_name='Vazir'
        ))
        self.exec_amt_input = TextInput(
            hint_text=fa("مبلغ را وارد کنید (ریال) - صفر مجاز است"),
            input_filter='float',
            size_hint=(1, None),
            height=dp(50),
            font_size=dp(14),
            multiline=False,
            padding=(dp(10), dp(12)),
            font_name='Vazir'
        )
        self.exec_amt_input.bind(text=lambda *x: self.update_fields_state())
        self.scroll_content.add_widget(self.exec_amt_input)

        exec_index_label_layout = BoxLayout(orientation='vertical', spacing=dp(2), size_hint=(1, None), height=dp(60))
        exec_index_label = Label(
            text=fa("شاخص دوره اجرا:"),
            font_size=dp(13),
            size_hint=(1, 0.4),
            font_name='Vazir'
        )
        exec_index_label_layout.add_widget(exec_index_label)
        exec_help_label = Label(
            text=fa(self.boq_index_info),
            font_size=dp(9),
            color=(0, 0.6, 0, 1),
            size_hint=(1, 0.6),
            text_size=(Window.width - dp(40), None),
            halign='left',
            valign='top',
            font_name='Vazir'
        )
        exec_index_label_layout.add_widget(exec_help_label)
        self.scroll_content.add_widget(exec_index_label_layout)

        self.exec_container = BoxLayout(orientation='vertical', size_hint=(1, None), spacing=dp(5))
        self.scroll_content.add_widget(self.exec_container)

        self.scroll_content.add_widget(Label(
            text=fa("شاخص مبنا اجرا:"),
            font_size=dp(13),
            size_hint=(1, None),
            height=dp(25),
            font_name='Vazir'
        ))
        self.exec_base_input = TextInput(
            hint_text=fa("شاخص مبنا را وارد کنید"),
            input_filter='float',
            size_hint=(1, None),
            height=dp(50),
            font_size=dp(14),
            multiline=False,
            padding=(dp(10), dp(12)),
            font_name='Vazir'
        )
        self.scroll_content.add_widget(self.exec_base_input)

        # ========== تجهیز کارگاه ==========
        work_title = Label(
            text=fa("اطلاعات تجهیز کارگاه"),
            font_size=dp(15),
            bold=True,
            size_hint=(1, None),
            height=dp(40),
            color=(0.3, 0.5, 0.2, 1),
            font_name='Vazir'
        )
        self.scroll_content.add_widget(work_title)

        self.scroll_content.add_widget(Label(
            text=fa("مبلغ تجهیز کارگاه:"),
            font_size=dp(13),
            size_hint=(1, None),
            height=dp(25),
            font_name='Vazir'
        ))
        self.work_amt_input = TextInput(
            hint_text=fa("مبلغ را وارد کنید (ریال) - صفر مجاز است"),
            input_filter='float',
            size_hint=(1, None),
            height=dp(50),
            font_size=dp(14),
            multiline=False,
            padding=(dp(10), dp(12)),
            font_name='Vazir'
        )
        self.work_amt_input.bind(text=lambda *x: self.update_fields_state())
        self.scroll_content.add_widget(self.work_amt_input)

        if self.app.boq_info.get('group') == "گروه ۴":
            help_text = "متوسط شاخص فصل ۳ ابنیه و ۳۵ مکانیک را وارد کنید"
        else:
            help_text = "میانگین شاخص رشته ای فهرست بهای منضم به پیمان و فهرست بهای واحد پایه ابنیه را وارد کنید"
        self.work_help_label = Label(
            text=fa(help_text),
            font_size=dp(9),
            color=(0, 0.6, 0, 1),
            size_hint=(1, None),
            text_size=(Window.width - dp(40), None),
            halign='left',
            valign='top',
            font_name='Vazir'
        )
        self.scroll_content.add_widget(self.work_help_label)

        work_index_label = Label(
            text=fa("شاخص دوره تجهیز کارگاه:"),
            font_size=dp(13),
            size_hint=(1, None),
            height=dp(25),
            font_name='Vazir'
        )
        self.scroll_content.add_widget(work_index_label)
        self.work_container = BoxLayout(orientation='vertical', size_hint=(1, None), spacing=dp(5))
        self.scroll_content.add_widget(self.work_container)

        self.scroll_content.add_widget(Label(
            text=fa("شاخص مبنا تجهیز کارگاه:"),
            font_size=dp(13),
            size_hint=(1, None),
            height=dp(25),
            font_name='Vazir'
        ))
        self.work_base_input = TextInput(
            hint_text=fa("شاخص مبنا را وارد کنید"),
            input_filter='float',
            size_hint=(1, None),
            height=dp(50),
            font_size=dp(14),
            multiline=False,
            padding=(dp(10), dp(12)),
            font_name='Vazir'
        )
        self.scroll_content.add_widget(self.work_base_input)

        # ========== روزها ==========
        days_title = Label(
            text=fa("روزهای مربوط به هر دوره (برای این صورت‌وضعیت)"),
            font_size=dp(13),
            bold=True,
            size_hint=(1, None),
            height=dp(40),
            color=(0, 0.6, 0, 1),
            font_name='Vazir'
        )
        self.scroll_content.add_widget(days_title)

        self.days_container = BoxLayout(orientation='vertical', size_hint=(1, None), spacing=dp(5))
        self.scroll_content.add_widget(self.days_container)

        self.scroll_content.add_widget(Label(text="", size_hint=(1, None), height=dp(10), font_name='Vazir'))

        scroll_view = ScrollView(size_hint=(1, 0.55))
        scroll_view.add_widget(self.scroll_content)
        self.add_widget(scroll_view)

        self.result_label = Label(
            text="",
            font_size=dp(12),
            size_hint=(1, 0.2),
            text_size=(Window.width - dp(20), None),
            halign='left',
            valign='top',
            color=(0, 0.6, 0, 1),
            font_name='Vazir'
        )
        self.add_widget(self.result_label)

        btn_layout = BoxLayout(spacing=dp(12), size_hint=(1, 0.12))

        self.calc_btn = Button(
            text=fa("محاسبه"),
            font_size=dp(14),
            background_color=(0.2, 0.6, 0.2, 1),
            size_hint=(0.5, None),
            height=dp(50),
            font_name='Vazir'
        )
        self.calc_btn.bind(on_press=self.calculate)
        btn_layout.add_widget(self.calc_btn)

        self.next_btn = Button(
            text=fa("صورت‌وضعیت بعدی"),
            font_size=dp(14),
            background_color=(0.3, 0.3, 0.8, 1),
            disabled=True,
            size_hint=(0.5, None),
            height=dp(50),
            font_name='Vazir'
        )
        self.next_btn.bind(on_press=self.next_statement)
        btn_layout.add_widget(self.next_btn)

        self.add_widget(btn_layout)

        self.on_period_change(self.period_spinner, '1')

    def on_period_change(self, spinner, text):
        count = int(text)
        self.rebuild_period_inputs(count)

    def rebuild_period_inputs(self, count):
        self.iem_container.clear_widgets()
        self.iew_container.clear_widgets()
        self.exec_container.clear_widgets()
        self.work_container.clear_widgets()
        self.days_container.clear_widgets()

        self.iem_inputs = []
        self.iew_inputs = []
        self.exec_inputs = []
        self.work_inputs = []
        self.days_inputs = []

        row_height = dp(40)

        # Iem
        iem_row = BoxLayout(orientation='horizontal', spacing=dp(8), size_hint=(1, None), height=row_height)
        for i in range(count):
            inp = TextInput(
                hint_text=fa(f"دوره {i+1}"),
                input_filter='float',
                size_hint=(1/count, 1),
                font_size=dp(12),
                multiline=False,
                padding=(dp(5), dp(8)),
                font_name='Vazir'
            )
            iem_row.add_widget(inp)
            self.iem_inputs.append(inp)
        self.iem_container.add_widget(iem_row)

        # Iew
        iew_row = BoxLayout(orientation='horizontal', spacing=dp(8), size_hint=(1, None), height=row_height)
        for i in range(count):
            inp = TextInput(
                hint_text=fa(f"دوره {i+1}"),
                input_filter='float',
                size_hint=(1/count, 1),
                font_size=dp(12),
                multiline=False,
                padding=(dp(5), dp(8)),
                font_name='Vazir'
            )
            iew_row.add_widget(inp)
            self.iew_inputs.append(inp)
        self.iew_container.add_widget(iew_row)

        # اجرا
        exec_row = BoxLayout(orientation='horizontal', spacing=dp(8), size_hint=(1, None), height=row_height)
        for i in range(count):
            inp = TextInput(
                hint_text=fa(f"دوره {i+1}"),
                input_filter='float',
                size_hint=(1/count, 1),
                font_size=dp(12),
                multiline=False,
                padding=(dp(5), dp(8)),
                font_name='Vazir'
            )
            exec_row.add_widget(inp)
            self.exec_inputs.append(inp)
        self.exec_container.add_widget(exec_row)

        # تجهیز
        work_row = BoxLayout(orientation='horizontal', spacing=dp(8), size_hint=(1, None), height=row_height)
        for i in range(count):
            inp = TextInput(
                hint_text=fa(f"دوره {i+1}"),
                input_filter='float',
                size_hint=(1/count, 1),
                font_size=dp(12),
                multiline=False,
                padding=(dp(5), dp(8)),
                font_name='Vazir'
            )
            work_row.add_widget(inp)
            self.work_inputs.append(inp)
        self.work_container.add_widget(work_row)

        # روزها
        days_row = BoxLayout(orientation='horizontal', spacing=dp(8), size_hint=(1, None), height=row_height)
        for i in range(count):
            inp = TextInput(
                text="۹۰",
                hint_text=fa(f"روزهای دوره {i+1}"),
                input_filter='int',
                size_hint=(1/count, 1),
                font_size=dp(12),
                multiline=False,
                padding=(dp(5), dp(8)),
                font_name='Vazir'
            )
            days_row.add_widget(inp)
            self.days_inputs.append(inp)
        self.days_container.add_widget(days_row)

        self.update_fields_state()

    def update_fields_state(self):
        """بر اساس مقادیر مبالغ، فیلدهای مربوطه را فعال/غیرفعال می‌کند"""
        count = int(self.period_spinner.text)

        # خرید
        buy_text = to_english_numbers(self.buy_amt_input.text)
        if buy_text and float(buy_text) == 0:
            for inp in self.iem_inputs + self.iew_inputs:
                inp.disabled = True
                inp.text = ""  # پاک کردن مقادیر
        else:
            for inp in self.iem_inputs + self.iew_inputs:
                inp.disabled = False

        # اجرا
        exec_text = to_english_numbers(self.exec_amt_input.text)
        if exec_text and float(exec_text) == 0:
            for inp in self.exec_inputs:
                inp.disabled = True
                inp.text = ""
        else:
            for inp in self.exec_inputs:
                inp.disabled = False

        # تجهیز
        work_text = to_english_numbers(self.work_amt_input.text)
        if work_text and float(work_text) == 0:
            for inp in self.work_inputs:
                inp.disabled = True
                inp.text = ""
        else:
            for inp in self.work_inputs:
                inp.disabled = False

    def validate_inputs(self):
        """بررسی اعتبار ورودی‌ها با توجه به صفر بودن مبالغ"""
        count = int(self.period_spinner.text)

        # مبالغ
        buy_amt = to_english_numbers(self.buy_amt_input.text)
        if not buy_amt:
            raise ValueError("مبلغ خرید نمی‌تواند خالی باشد")
        buy_amt = float(buy_amt)
        if buy_amt < 0:
            raise ValueError("مبلغ خرید نمی‌تواند منفی باشد")

        exec_amt = to_english_numbers(self.exec_amt_input.text)
        if not exec_amt:
            raise ValueError("مبلغ اجرا نمی‌تواند خالی باشد")
        exec_amt = float(exec_amt)
        if exec_amt < 0:
            raise ValueError("مبلغ اجرا نمی‌تواند منفی باشد")

        work_amt = to_english_numbers(self.work_amt_input.text)
        if not work_amt:
            raise ValueError("مبلغ تجهیز کارگاه نمی‌تواند خالی باشد")
        work_amt = float(work_amt)
        if work_amt < 0:
            raise ValueError("مبلغ تجهیز کارگاه نمی‌تواند منفی باشد")

        # روزها
        days = []
        for inp in self.days_inputs:
            d_text = to_english_numbers(inp.text)
            if not d_text or int(d_text) <= 0:
                raise ValueError("تمام روزهای دوره‌ها باید عددی مثبت باشند")
            days.append(int(d_text))

        # شاخص‌های مبنا (در صورت اول)
        if self.current_statement == 1:
            ibm = to_english_numbers(self.Ibm_input.text)
            ibw = to_english_numbers(self.Ibw_input.text)
            exec_base = to_english_numbers(self.exec_base_input.text)
            work_base = to_english_numbers(self.work_base_input.text)
            if not ibm or not ibw or not exec_base or not work_base:
                raise ValueError("تمام شاخص‌های مبنا باید پر شوند")
            if float(ibm) <= 0 or float(ibw) <= 0 or float(exec_base) <= 0 or float(work_base) <= 0:
                raise ValueError("شاخص‌های مبنا باید بزرگتر از صفر باشند")

        # بررسی شاخص‌های خرید (اگر مبلغ خرید > 0)
        if buy_amt > 0:
            for i in range(count):
                if days[i] > 0:
                    iem = to_english_numbers(self.iem_inputs[i].text)
                    iew = to_english_numbers(self.iew_inputs[i].text)
                    if not iem or not iew:
                        raise ValueError(f"دوره {i+1}: شاخص خرید و شاخص دستمزدی خرید باید پر شوند")
                    if float(iem) <= 0 or float(iew) <= 0:
                        raise ValueError(f"دوره {i+1}: شاخص‌های خرید باید بزرگتر از صفر باشند")

        # بررسی شاخص‌های اجرا (اگر مبلغ اجرا > 0)
        if exec_amt > 0:
            for i in range(count):
                if days[i] > 0:
                    exec_idx = to_english_numbers(self.exec_inputs[i].text)
                    if not exec_idx:
                        raise ValueError(f"دوره {i+1}: شاخص اجرا باید پر شود")
                    if float(exec_idx) <= 0:
                        raise ValueError(f"دوره {i+1}: شاخص اجرا باید بزرگتر از صفر باشد")

        # بررسی شاخص‌های تجهیز (اگر مبلغ تجهیز > 0)
        if work_amt > 0:
            for i in range(count):
                if days[i] > 0:
                    work_idx = to_english_numbers(self.work_inputs[i].text)
                    if not work_idx:
                        raise ValueError(f"دوره {i+1}: شاخص تجهیز کارگاه باید پر شود")
                    if float(work_idx) <= 0:
                        raise ValueError(f"دوره {i+1}: شاخص تجهیز کارگاه باید بزرگتر از صفر باشد")

        return True

    def calculate(self, instance):
        try:
            self.validate_inputs()

            buy_amt = float(to_english_numbers(self.buy_amt_input.text))
            q = float(self.q_input.text)
            exec_amt = float(to_english_numbers(self.exec_amt_input.text))
            work_amt = float(to_english_numbers(self.work_amt_input.text))

            count = int(self.period_spinner.text)
            days = []
            total_days_this = 0
            for inp in self.days_inputs:
                d = int(to_english_numbers(inp.text))
                days.append(d)
                total_days_this += d

            # شاخص‌های مبنا (صورت اول)
            if self.current_statement == 1:
                self.Ibm_base = float(to_english_numbers(self.Ibm_input.text))
                self.Ibw_base = float(to_english_numbers(self.Ibw_input.text))
                self.exec_base_value = float(to_english_numbers(self.exec_base_input.text))
                self.work_base_value = float(to_english_numbers(self.work_base_input.text))

                for inp in [self.Ibm_input, self.Ibw_input, self.exec_base_input, self.work_base_input]:
                    inp.readonly = True

                self.app.base_indices = {
                    'Ibm': self.Ibm_base,
                    'Ibw': self.Ibw_base,
                    'exec_base': self.exec_base_value,
                    'work_base': self.work_base_value
                }

            # ========== محاسبات خرید ==========
            if buy_amt > 0:
                iem_values = [float(to_english_numbers(inp.text)) for inp in self.iem_inputs]
                iew_values = [float(to_english_numbers(inp.text)) for inp in self.iew_inputs]
                alpha_buy_sum = 0
                days_used_buy = 0
                alpha_buy_periods = []
                for i in range(count):
                    if days[i] == 0:
                        alpha_buy_periods.append(None)
                        continue
                    alpha = calc_alpha_buy(q, iem_values[i], self.Ibm_base, iew_values[i], self.Ibw_base, self.app.delay_factor)
                    alpha_buy_sum += alpha * days[i]
                    days_used_buy += days[i]
                    alpha_buy_periods.append(alpha)
                alpha_buy_current = alpha_buy_sum / days_used_buy
                buy_adj = buy_amt * alpha_buy_current
            else:
                iem_values = [0]*count
                iew_values = [0]*count
                alpha_buy_current = 0
                alpha_buy_periods = [None]*count
                buy_adj = 0

            # ========== محاسبات اجرا ==========
            if exec_amt > 0:
                exec_idx_values = [float(to_english_numbers(inp.text)) for inp in self.exec_inputs]
                alpha_exec_sum = 0
                days_used_exec = 0
                alpha_exec_periods = []
                for i in range(count):
                    if days[i] == 0:
                        alpha_exec_periods.append(None)
                        continue
                    alpha = calc_alpha_simple(exec_idx_values[i], self.exec_base_value, self.app.delay_factor)
                    alpha_exec_sum += alpha * days[i]
                    days_used_exec += days[i]
                    alpha_exec_periods.append(alpha)
                alpha_exec_current = alpha_exec_sum / days_used_exec
                exec_adj = exec_amt * alpha_exec_current
            else:
                exec_idx_values = [0]*count
                alpha_exec_current = 0
                alpha_exec_periods = [None]*count
                exec_adj = 0

            # ========== محاسبات تجهیز ==========
            if work_amt > 0:
                work_idx_values = [float(to_english_numbers(inp.text)) for inp in self.work_inputs]
                alpha_work_sum = 0
                days_used_work = 0
                alpha_work_periods = []
                for i in range(count):
                    if days[i] == 0:
                        alpha_work_periods.append(None)
                        continue
                    alpha = calc_alpha_simple(work_idx_values[i], self.work_base_value, self.app.delay_factor)
                    alpha_work_sum += alpha * days[i]
                    days_used_work += days[i]
                    alpha_work_periods.append(alpha)
                alpha_work_current = alpha_work_sum / days_used_work
                work_adj = work_amt * alpha_work_current
            else:
                work_idx_values = [0]*count
                alpha_work_current = 0
                alpha_work_periods = [None]*count
                work_adj = 0

            total_adj = buy_adj + exec_adj + work_adj
            self.total_adjustment += total_adj

            detail = {
                'statement': self.current_statement,
                'buy_amt': buy_amt,
                'exec_amt': exec_amt,
                'work_amt': work_amt,
                'periods': count,
                'days': days,
                'iem': iem_values,
                'iew': iew_values,
                'exec_idx': exec_idx_values,
                'work_idx': work_idx_values,
                'alpha_buy_periods': alpha_buy_periods,
                'alpha_exec_periods': alpha_exec_periods,
                'alpha_work_periods': alpha_work_periods,
                'alpha_buy': alpha_buy_current,
                'alpha_exec': alpha_exec_current,
                'alpha_work': alpha_work_current,
                'buy_adj': buy_adj,
                'exec_adj': exec_adj,
                'work_adj': work_adj,
                'total_adj': total_adj,
                'cumulative': self.total_adjustment
            }
            self.app.statement_details.append(detail)

            self.app.statement_results.append({
                'statement': self.current_statement,
                'buy_adj': buy_adj,
                'exec_adj': exec_adj,
                'work_adj': work_adj,
                'total_adj': total_adj,
                'cumulative': self.total_adjustment
            })

            if self.total_days_so_far == 0:
                self.prev_alpha_buy = alpha_buy_current
                self.prev_alpha_exec = alpha_exec_current
                self.prev_alpha_work = alpha_work_current
            else:
                self.prev_alpha_buy = (self.prev_alpha_buy * self.total_days_so_far + alpha_buy_current * total_days_this) / (self.total_days_so_far + total_days_this)
                self.prev_alpha_exec = (self.prev_alpha_exec * self.total_days_so_far + alpha_exec_current * total_days_this) / (self.total_days_so_far + total_days_this)
                self.prev_alpha_work = (self.prev_alpha_work * self.total_days_so_far + alpha_work_current * total_days_this) / (self.total_days_so_far + total_days_this)
            self.total_days_so_far += total_days_this

            result_text = fa(
                f"نتایج صورت‌وضعیت {self.current_statement}:\n\n"
                f"تعدیل خرید: {format_persian_number(buy_adj)}\n"
                f"تعدیل اجرا: {format_persian_number(exec_adj)}\n"
                f"تعدیل تجهیز کارگاه: {format_persian_number(work_adj)}\n"
                f"جمع صورت‌وضعیت: {format_persian_number(total_adj)}\n"
                f"جمع کل تعدیل: {format_persian_number(self.total_adjustment)}"
            )
            self.result_label.text = result_text

            self.calc_btn.disabled = True
            self.next_btn.disabled = False

        except Exception as e:
            popup = Popup(
                title=fa('خطا'),
                title_font='Vazir',
                content=Label(text=fa(str(e)), font_name='Vazir'),
                size_hint=(0.8, 0.4)
            )
            popup.open()

    def next_statement(self, instance):
        if self.current_statement < self.app.statement_count:
            self.current_statement += 1
            self.header.text = fa(f"صورت‌وضعیت {self.current_statement}")

            # پاک کردن فیلدهای مبلغ و شاخص‌ها
            for inp in self.iem_inputs + self.iew_inputs + self.exec_inputs + self.work_inputs:
                inp.text = ""
            count = int(self.period_spinner.text)
            for inp in self.days_inputs:
                inp.text = "۹۰"
                inp.disabled = False

            self.buy_amt_input.text = ""
            self.exec_amt_input.text = ""
            self.work_amt_input.text = ""

            self.calc_btn.disabled = False
            self.next_btn.disabled = True
            self.result_label.text = ""

            self.update_fields_state()  # تنظیم وضعیت فیلدها بر اساس مبالغ (که خالی هستند)
        else:
            self.app.show_final_results()

# ================== صفحه نتایج نهایی و خروجی اکسل ==================
class FinalResultsScreen(BoxLayout):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs)
        self.app = app
        self.orientation = 'vertical'
        self.padding = dp(10)
        self.spacing = dp(10)

        title = Label(
            text=fa("نتایج نهایی تعدیل"),
            font_size=dp(18),
            bold=True,
            size_hint=(1, 0.1),
            font_name='Vazir'
        )
        self.add_widget(title)

        col_width = (Window.width - dp(20)) / 5

        grid = GridLayout(cols=5, spacing=dp(2), size_hint_y=None, row_force_default=False)
        grid.bind(minimum_height=grid.setter('height'))

        headers = ["شماره", "تعدیل خرید", "تعدیل اجرا", "تعدیل تجهیز", "جمع صورت‌وضعیت"]
        for h in headers:
            btn = Button(
                text=fa(h),
                bold=True,
                size_hint_y=None,
                height=dp(40),
                font_name='Vazir',
                font_size=dp(12),
                color=(0,0,0,1),
                background_normal='',
                background_color=(0.8,0.8,0.8,1),
                disabled=True,
                text_size=(col_width, None),
                halign='center',
                valign='middle'
            )
            grid.add_widget(btn)

        for res in self.app.statement_results:
            lbl = Label(
                text=fa(to_persian_numbers(str(res['statement']))),
                font_name='Vazir',
                font_size=dp(11),
                size_hint_y=None,
                height=dp(30),
                text_size=(col_width, None),
                halign='center',
                valign='middle'
            )
            grid.add_widget(lbl)
            lbl = Label(
                text=format_persian_number(res['buy_adj']),
                font_name='Vazir',
                font_size=dp(11),
                size_hint_y=None,
                height=dp(30),
                text_size=(col_width, None),
                halign='center',
                valign='middle'
            )
            grid.add_widget(lbl)
            lbl = Label(
                text=format_persian_number(res['exec_adj']),
                font_name='Vazir',
                font_size=dp(11),
                size_hint_y=None,
                height=dp(30),
                text_size=(col_width, None),
                halign='center',
                valign='middle'
            )
            grid.add_widget(lbl)
            lbl = Label(
                text=format_persian_number(res['work_adj']),
                font_name='Vazir',
                font_size=dp(11),
                size_hint_y=None,
                height=dp(30),
                text_size=(col_width, None),
                halign='center',
                valign='middle'
            )
            grid.add_widget(lbl)
            lbl = Label(
                text=format_persian_number(res['total_adj']),
                font_name='Vazir',
                font_size=dp(11),
                size_hint_y=None,
                height=dp(30),
                text_size=(col_width, None),
                halign='center',
                valign='middle'
            )
            grid.add_widget(lbl)

        total = self.app.statement_results[-1]['cumulative'] if self.app.statement_results else 0
        lbl = Label(text=fa("جمع کل"), bold=True, font_name='Vazir', font_size=dp(12), size_hint_y=None, height=dp(30), text_size=(col_width, None), halign='center', valign='middle')
        grid.add_widget(lbl)
        grid.add_widget(Label(text="", font_name='Vazir', size_hint_y=None, height=dp(30)))
        grid.add_widget(Label(text="", font_name='Vazir', size_hint_y=None, height=dp(30)))
        grid.add_widget(Label(text="", font_name='Vazir', size_hint_y=None, height=dp(30)))
        lbl_total = Label(text=format_persian_number(total), bold=True, font_name='Vazir', font_size=dp(12), size_hint_y=None, height=dp(30), text_size=(col_width, None), halign='center', valign='middle')
        grid.add_widget(lbl_total)

        scroll = ScrollView(size_hint=(1, 0.7))
        scroll.add_widget(grid)
        self.add_widget(scroll)

        btn_layout = BoxLayout(size_hint=(1, 0.15), spacing=dp(15))

        export_btn = Button(
            text=fa("خروجی اکسل"),
            font_size=dp(14),
            background_color=(0.2, 0.6, 0.2, 1),
            size_hint=(0.5, None),
            height=dp(50),
            font_name='Vazir'
        )
        export_btn.bind(on_press=self.export_excel)
        btn_layout.add_widget(export_btn)

        back_btn = Button(
            text=fa("شروع جدید"),
            font_size=dp(14),
            background_color=(0.8, 0.3, 0.3, 1),
            size_hint=(0.5, None),
            height=dp(50),
            font_name='Vazir'
        )
        back_btn.bind(on_press=lambda x: self.app.show_start_screen())
        btn_layout.add_widget(back_btn)

        self.add_widget(btn_layout)

    def export_excel(self, instance):
        if not OPENPYXL_AVAILABLE:
            popup = Popup(
                title=fa('خطا'),
                title_font='Vazir',
                content=Label(text=fa("کتابخانه openpyxl نصب نیست.\nنصب: pip install openpyxl"), font_name='Vazir'),
                size_hint=(0.8, 0.4)
            )
            popup.open()
            return

        self._save_excel_fallback()

    def _save_excel_fallback(self):
        try:
            download_dir = '/storage/emulated/0/Download'
            if not os.path.exists(download_dir):
                download_dir = os.path.join(os.environ.get('EXTERNAL_STORAGE', '/sdcard'), 'Download')
            os.makedirs(download_dir, exist_ok=True)
            filename = os.path.join(download_dir, "نتایج_تعدیل.xlsx")
            self._save_excel_to_path(filename)
            popup = Popup(
                title=fa('موفقیت'),
                title_font='Vazir',
                content=Label(text=fa(f"فایل با موفقیت در مسیر زیر ذخیره شد:\n{filename}"), font_name='Vazir'),
                size_hint=(0.8, 0.4)
            )
            popup.open()
        except Exception as e:
            popup = Popup(
                title=fa('خطا در ذخیره'),
                title_font='Vazir',
                content=Label(text=fa(str(e)), font_name='Vazir'),
                size_hint=(0.8, 0.4)
            )
            popup.open()

    def _save_excel_to_path(self, path):
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_info = wb.create_sheet("اطلاعات پروژه")
        ws_info.column_dimensions['A'].width = 25
        ws_info.column_dimensions['B'].width = 60

        title_cell = ws_info.cell(row=1, column=1, value="مشخصه")
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell = ws_info.cell(row=1, column=2, value="مقدار")
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        row = 2
        ws_info.cell(row=row, column=1, value="گروه ")
        ws_info.cell(row=row, column=2, value=self.app.boq_info.get('group', ''))
        row += 1
        if self.app.boq_info.get('subgroup'):
            ws_info.cell(row=row, column=1, value="زیرگروه ")
            ws_info.cell(row=row, column=2, value=self.app.boq_info.get('subgroup', ''))
            row += 1
        ws_info.cell(row=row, column=1, value="شاخص متناظر ")
        ws_info.cell(row=row, column=2, value=self.app.boq_info.get('value', ''))
        row += 1

        ws_info.cell(row=row, column=1, value="گروه کالایی")
        ws_info.cell(row=row, column=2, value=self.app.q_info.get('group', ''))
        row += 1
        ws_info.cell(row=row, column=1, value="ردیف کالایی")
        ws_info.cell(row=row, column=2, value=self.app.q_info.get('row', ''))
        row += 1
        ws_info.cell(row=row, column=1, value="توضیحات ")
        ws_info.cell(row=row, column=2, value=Q_DESC.get(self.app.q_info.get('row', ''), ''))
        row += 1
        ws_info.cell(row=row, column=1, value="ضریب q")
        ws_info.cell(row=row, column=2, value=self.app.q_info.get('value', ''))
        row += 1
        ws_info.cell(row=row, column=1, value="اطلاعات ردیف ")
        ws_info.cell(row=row, column=2, value=self.app.q_info.get('row_info', ''))
        row += 1

        ws_info.cell(row=row, column=1, value="ضریب تأخیر")
        ws_info.cell(row=row, column=2, value=self.app.delay_factor)
        row += 1

        if self.app.base_indices:
            ws_info.cell(row=row, column=1, value="شاخص مبنا خرید ")
            ws_info.cell(row=row, column=2, value=self.app.base_indices.get('Ibm', ''))
            row += 1
            ws_info.cell(row=row, column=1, value="شاخص مبنای دستمزدی خرید ")
            ws_info.cell(row=row, column=2, value=self.app.base_indices.get('Ibw', ''))
            row += 1
            ws_info.cell(row=row, column=1, value="شاخص مبنا اجرا")
            ws_info.cell(row=row, column=2, value=self.app.base_indices.get('exec_base', ''))
            row += 1
            ws_info.cell(row=row, column=1, value="شاخص مبنا تجهیز کارگاه")
            ws_info.cell(row=row, column=2, value=self.app.base_indices.get('work_base', ''))
            row += 1

        ws_periods = wb.create_sheet("جزئیات دوره‌ها")
        headers_periods = [
            "شماره صورت‌وضعیت", "شماره دوره", "روزهای دوره",
            "شاخص خرید", "شاخص دستمزدی خرید", "آلفای خرید (دوره)",
            "شاخص اجرا", "آلفای اجرا (دوره)",
            "شاخص تجهیز", "آلفای تجهیز (دوره)"
        ]
        for col, h in enumerate(headers_periods, 1):
            cell = ws_periods.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            ws_periods.column_dimensions[get_column_letter(col)].width = 16

        row = 2
        for detail in self.app.statement_details:
            stmt = detail['statement']
            periods = detail['periods']
            for p in range(periods):
                ws_periods.cell(row=row, column=1, value=stmt)
                ws_periods.cell(row=row, column=2, value=p+1)
                ws_periods.cell(row=row, column=3, value=detail['days'][p])
                ws_periods.cell(row=row, column=4, value=detail['iem'][p])
                ws_periods.cell(row=row, column=5, value=detail['iew'][p])
                ws_periods.cell(row=row, column=6, value=detail['alpha_buy_periods'][p] if detail['alpha_buy_periods'][p] is not None else "")
                ws_periods.cell(row=row, column=7, value=detail['exec_idx'][p])
                ws_periods.cell(row=row, column=8, value=detail['alpha_exec_periods'][p] if detail['alpha_exec_periods'][p] is not None else "")
                ws_periods.cell(row=row, column=9, value=detail['work_idx'][p])
                ws_periods.cell(row=row, column=10, value=detail['alpha_work_periods'][p] if detail['alpha_work_periods'][p] is not None else "")
                row += 1

        ws_summary = wb.create_sheet("خلاصه صورت‌وضعیت‌ها")
        headers_summary = [
            "شماره", "مبلغ خرید", "مبلغ اجرا", "مبلغ تجهیز",
            "آلفای خرید", "آلفای اجرا", "آلفای تجهیز",
            "تعدیل خرید", "تعدیل اجرا", "تعدیل تجهیز",
            "جمع صورت‌وضعیت", "جمع تجمعی"
        ]
        for col, h in enumerate(headers_summary, 1):
            cell = ws_summary.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            ws_summary.column_dimensions[get_column_letter(col)].width = 14

        row = 2
        for detail in self.app.statement_details:
            ws_summary.cell(row=row, column=1, value=detail['statement'])
            ws_summary.cell(row=row, column=2, value=detail['buy_amt'])
            ws_summary.cell(row=row, column=3, value=detail['exec_amt'])
            ws_summary.cell(row=row, column=4, value=detail['work_amt'])
            ws_summary.cell(row=row, column=5, value=detail['alpha_buy'])
            ws_summary.cell(row=row, column=6, value=detail['alpha_exec'])
            ws_summary.cell(row=row, column=7, value=detail['alpha_work'])
            ws_summary.cell(row=row, column=8, value=detail['buy_adj'])
            ws_summary.cell(row=row, column=9, value=detail['exec_adj'])
            ws_summary.cell(row=row, column=10, value=detail['work_adj'])
            ws_summary.cell(row=row, column=11, value=detail['total_adj'])
            ws_summary.cell(row=row, column=12, value=detail['cumulative'])
            row += 1

        wb.save(path)

# ================== برنامه اصلی ==================
class SVCalculatorApp(App):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.statement_count = 0
        self.statement_results = []
        self.statement_details = []
        self.base_indices = {}
        self.boq_info = None
        self.q_info = None
        self.delay_factor = None
        self.current_screen = None

    def build(self):
        Window.softinput_mode = 'below_target'
        self.root_layout = BoxLayout(orientation='vertical')
        Clock.schedule_once(self.check_persian_libs, 0.5)
        self.show_start_screen()
        return self.root_layout

    def check_persian_libs(self, dt):
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
        except ImportError:
            popup = Popup(
                title='اخطار',
                title_font='Vazir',
                content=Label(
                    text='برای نمایش بهتر فارسی، کتابخانه‌های arabic_reshaper و python-bidi باید نصب شوند.\nممکن است متون به درستی نمایش داده نشوند.',
                    font_name='Vazir'
                ),
                size_hint=(0.8, 0.4)
            )
            popup.open()

    def show_start_screen(self):
        self.clear_screen()
        self.current_screen = StartScreen(self)
        self.root_layout.add_widget(self.current_screen)

    def show_boq_selection(self):
        self.clear_screen()
        self.current_screen = BOQSelectionScreen(self)
        self.root_layout.add_widget(self.current_screen)

    def show_q_selection(self):
        self.clear_screen()
        self.current_screen = QSelectionScreen(self)
        self.root_layout.add_widget(self.current_screen)

    def show_delay_selection(self):
        self.clear_screen()
        self.current_screen = DelaySelectionScreen(self)
        self.root_layout.add_widget(self.current_screen)

    def show_calculation_screen(self):
        self.clear_screen()
        self.current_screen = CalculationScreen(self)
        self.root_layout.add_widget(self.current_screen)

    def show_final_results(self):
        self.clear_screen()
        self.current_screen = FinalResultsScreen(self)
        self.root_layout.add_widget(self.current_screen)

    def clear_screen(self):
        self.root_layout.clear_widgets()

if __name__ == '__main__':
    SVCalculatorApp().run()